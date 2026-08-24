"""Per-pharmacy product catalog (OTC + parapharmacy) — the basis of the order/delivery circuit.
Gated by the `order_delivery` module; the pharmacist manages it manually or via XML import."""

from __future__ import annotations

import json
from datetime import datetime

from fastapi import APIRouter, Depends, File, Form, HTTPException, Query, Response, UploadFile, status
from pydantic import BaseModel, Field

from app.core.deps import TenantContext, require
from app.repositories.pharmacy_catalog import PharmacyCatalogRepository, StockMovementRepository
from app.repositories.pharmacy_categories import PharmacyCategoryRepository
from app.repositories.shop_campaigns import ShopCampaignRepository
from app.repositories.shop_order_discounts import ShopOrderDiscountRepository
from app.repositories.shop_promos import ShopBundleRepository, ShopCouponRepository
from app.repositories.shop_service_offers import ShopServiceOffersRepository
from app.repositories.supplier_settings import SupplierSettingsRepository

_MAX_XML = 25 * 1024 * 1024  # 25 MB cap on the uploaded catalog XML (no unbounded in-memory read)
_XML_CTYPES = {"application/xml", "text/xml", "application/octet-stream", ""}
_MAX_IMG = 8 * 1024 * 1024   # 8 MB cap on a product photo upload

router = APIRouter()
_MODULE = "order_delivery"      # opt-in module — ενεργοποιείται ανά φαρμακείο
_PERM = "portal:manage"


def _repo(ctx: TenantContext) -> PharmacyCatalogRepository:
    return PharmacyCatalogRepository(tenant_id=ctx.tenant_id)


@router.get("")
async def list_products(q: str = "", category: str | None = None, type: str | None = None,
                        tag: str | None = None, sort: str = "featured", in_stock: bool = False,
                        for_sale: bool = False, page: int = 1,
                        ctx: TenantContext = Depends(require(_PERM, module=_MODULE))):
    return await _repo(ctx).list(q=q, category=category, ptype=type, tag=tag, sort=sort,
                                 in_stock_only=in_stock, for_sale_only=for_sale, page=page)


@router.get("/categories")
async def categories(ctx: TenantContext = Depends(require(_PERM, module=_MODULE))):
    return {"categories": await _repo(ctx).categories(), "tags": await _repo(ctx).tags()}


# ── Δέντρο κατηγοριών e-shop (3 επίπεδα) ────────────────────────────────────
def _catrepo(ctx: TenantContext) -> PharmacyCategoryRepository:
    return PharmacyCategoryRepository(tenant_id=ctx.tenant_id)


@router.get("/category-tree")
async def category_tree(ctx: TenantContext = Depends(require(_PERM, module=_MODULE))):
    return {"items": await _catrepo(ctx).tree()}


class CategoryIn(BaseModel):
    name: str
    parent_id: str | None = None


@router.post("/category")
async def add_category(body: CategoryIn, ctx: TenantContext = Depends(require(_PERM, module=_MODULE))):
    return await _catrepo(ctx).add(body.name, body.parent_id)


@router.post("/category/seed-from-products")
async def seed_categories(ctx: TenantContext = Depends(require(_PERM, module=_MODULE))):
    """Μεταφορά των υπαρχουσών κατηγοριών ειδών στο δέντρο ως Κατηγορία 1 + ανάθεση στα είδη."""
    return await _catrepo(ctx).seed_from_products()


class CategoryRenameIn(BaseModel):
    name: str


@router.put("/category/{cat_id}")
async def rename_category(cat_id: str, body: CategoryRenameIn,
                          ctx: TenantContext = Depends(require(_PERM, module=_MODULE))):
    return await _catrepo(ctx).rename(cat_id, body.name)


class CategoryIconIn(BaseModel):
    icon: str | None = None


@router.post("/category/{cat_id}/icon")
async def set_category_icon(cat_id: str, body: CategoryIconIn,
                            ctx: TenantContext = Depends(require(_PERM, module=_MODULE))):
    return await _catrepo(ctx).set_icon(cat_id, body.icon)


@router.delete("/category/{cat_id}")
async def delete_category(cat_id: str, ctx: TenantContext = Depends(require(_PERM, module=_MODULE))):
    return await _catrepo(ctx).delete(cat_id)


@router.post("/category/{cat_id}/image")
async def category_image(cat_id: str, file: UploadFile = File(...),
                         ctx: TenantContext = Depends(require(_PERM, module=_MODULE))):
    """Φωτογραφία κατηγορίας (για το μενού-πλακίδια της πύλης). Reuse του image storage/serve των ειδών."""
    content = await file.read(_MAX_IMG + 1)
    if len(content) > _MAX_IMG:
        raise HTTPException(status.HTTP_413_REQUEST_ENTITY_TOO_LARGE, detail={"error": "file_too_large"})
    image_id = await _repo(ctx).save_image(content, (file.content_type or "").split(";")[0].strip().lower())
    if not image_id:
        raise HTTPException(status.HTTP_400_BAD_REQUEST, detail={"error": "bad_image"})
    await _catrepo(ctx).set_image(cat_id, image_id)
    return {"ok": True, "image_id": image_id}


@router.delete("/category/{cat_id}/image")
async def category_image_remove(cat_id: str, ctx: TenantContext = Depends(require(_PERM, module=_MODULE))):
    return await _catrepo(ctx).set_image(cat_id, None)


def _read_grid(content: bytes, filename: str) -> list[list]:
    """Διαβάζει Excel(.xlsx)/CSV σε 2D λίστα γραμμών (ευέλικτο: ο χρήστης ορίζει μετά στήλες/γραμμή)."""
    import io
    name = (filename or "").lower()
    if name.endswith((".csv", ".txt", ".tsv")):
        import csv
        text = content.decode("utf-8-sig", errors="replace")
        sample = text[:4000]
        delim = ";" if sample.count(";") > sample.count(",") else ("\t" if sample.count("\t") > sample.count(",") else ",")
        return [list(r) for r in csv.reader(io.StringIO(text), delimiter=delim)]
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    rows = [list(r) for r in ws.iter_rows(values_only=True)]
    wb.close()
    return rows


@router.post("/category/import/preview")
async def category_import_preview(file: UploadFile = File(...),
                                  ctx: TenantContext = Depends(require(_PERM, module=_MODULE))):
    """Επιστρέφει τις πρώτες γραμμές ώστε ο χρήστης να διαλέξει γραμμή-έναρξης & στήλες Κατ.1/2/3."""
    content = await file.read(_MAX_XML + 1)
    if len(content) > _MAX_XML:
        raise HTTPException(status.HTTP_413_REQUEST_ENTITY_TOO_LARGE, detail={"error": "file_too_large"})
    try:
        grid = _read_grid(content, file.filename or "")
    except Exception:  # noqa: BLE001
        raise HTTPException(status.HTTP_400_BAD_REQUEST, detail={"error": "parse_failed"})
    ncols = max((len(r) for r in grid[:30]), default=0)
    preview = [[("" if (i >= len(r) or r[i] is None) else str(r[i]))[:60] for i in range(ncols)] for r in grid[:12]]
    return {"rows": preview, "ncols": ncols, "total_rows": len(grid)}


@router.post("/category/import")
async def category_import(file: UploadFile = File(...), start_row: int = Form(1),
                          col1: int = Form(0), col2: int = Form(-1), col3: int = Form(-1),
                          ctx: TenantContext = Depends(require(_PERM, module=_MODULE))):
    content = await file.read(_MAX_XML + 1)
    if len(content) > _MAX_XML:
        raise HTTPException(status.HTTP_413_REQUEST_ENTITY_TOO_LARGE, detail={"error": "file_too_large"})
    try:
        grid = _read_grid(content, file.filename or "")
    except Exception:  # noqa: BLE001
        raise HTTPException(status.HTTP_400_BAD_REQUEST, detail={"error": "parse_failed"})
    return await _catrepo(ctx).import_rows(grid, col1=col1,
                                           col2=col2 if col2 >= 0 else None,
                                           col3=col3 if col3 >= 0 else None,
                                           start_row=start_row)


@router.get("/taxonomy")
async def taxonomy(ctx: TenantContext = Depends(require(_PERM, module=_MODULE))):
    """Ενιαία ταξινομία (3 κλάσεις + κατηγορίες) — κοινή για όλα τα φαρμακεία."""
    from app.services.catalog_taxonomy import TAXONOMY
    return TAXONOMY


@router.get("/registry")
async def registry(q: str = "", category: str | None = None, page: int = 1,
                   ctx: TenantContext = Depends(require(_PERM, module=_MODULE))):
    """Μητρώο φαρμάκων ΗΔΥΚΑ (search/πλοήγηση ανά κατηγορία) — για ενεργοποίηση «προς πώληση»."""
    return await _repo(ctx).registry(q=q, category=category, page=page)


class ActivateIn(BaseModel):
    barcodes: list[str] | None = None
    category: str | None = None
    type: str = "rx_medicine"
    stock_qty: int = Field(0, ge=0)


@router.post("/activate")
async def activate(body: ActivateIn, ctx: TenantContext = Depends(require(_PERM, module=_MODULE))):
    return await _repo(ctx).activate(barcodes=body.barcodes, category=body.category,
                                     ptype=body.type, stock_qty=body.stock_qty)


@router.post("/image")
async def upload_image(file: UploadFile = File(...),
                       ctx: TenantContext = Depends(require(_PERM, module=_MODULE))):
    """Ανέβασμα φωτογραφίας προϊόντος → resize + αποθήκευση στη shared DB. Επιστρέφει image_id."""
    if not (file.content_type or "").lower().startswith("image/"):
        raise HTTPException(status.HTTP_415_UNSUPPORTED_MEDIA_TYPE, detail={"error": "not_an_image"})
    raw = await file.read(_MAX_IMG + 1)
    if len(raw) > _MAX_IMG:
        raise HTTPException(status.HTTP_413_REQUEST_ENTITY_TOO_LARGE, detail={"error": "file_too_large"})
    image_id = await _repo(ctx).save_image(raw, file.content_type or "")
    if not image_id:
        return {"ok": False, "error": "bad_image"}
    return {"ok": True, "image_id": image_id, "url": f"/catalog/image/{image_id}"}


@router.get("/image/{image_id}")
async def get_image(image_id: str):
    """ΔΗΜΟΣΙΟ (χωρίς auth) — τα <img> των πελατών/φαρμακείου δεν στέλνουν token· opaque id, μη-PII."""
    res = await PharmacyCatalogRepository.get_image(image_id)
    if not res:
        raise HTTPException(status.HTTP_404_NOT_FOUND)
    data, ctype = res
    return Response(content=data, media_type=ctype,
                    headers={"Cache-Control": "public, max-age=31536000, immutable"})


@router.get("/prefill")
async def prefill(barcode: str = Query(...),
                  ctx: TenantContext = Depends(require(_PERM, module=_MODULE))):
    return await _repo(ctx).prefill(barcode)


class ProductIn(BaseModel):
    barcode: str
    name: str
    description_short: str | None = None
    description_long: str | None = None
    photo_url: str | None = None
    price_cents: int = Field(0, ge=0)
    wholesale_cents: int = Field(0, ge=0)   # χονδρική — για την κερδοφορία
    vat_rate: int = Field(6, ge=0, le=30)          # συντελεστής ΦΠΑ % (6/13/24 στην Ελλάδα)
    price_includes_vat: bool = True                # αν η ΛΙΑΝΙΚΗ τιμή περιλαμβάνει ήδη ΦΠΑ
    type: str = "parapharmacy"          # rx_medicine | otc_medicine | parapharmacy
    category: str | None = None
    tags: list[str] = Field(default_factory=list)
    featured: bool = False
    image_id: str | None = None
    images: list[str] = Field(default_factory=list)   # gallery (image_id list)· κύρια = image_id/images[0]
    usage_video_url: str | None = None      # οδηγίες χρήσης (YouTube/Vimeo) — ο πελάτης το βλέπει
    discount_pct: int = Field(0, ge=0, le=90)
    stock_qty: int = Field(0, ge=0)
    active: bool = True                  # ενεργό/ανενεργό είδος στην αποθήκη
    for_sale: bool = False              # πωλείται στο e-shop → εμφανίζεται στον Κατάλογο
    # πλήρη χαρακτηριστικά αποθήκης
    min_stock: int = Field(0, ge=0)     # σημείο αναπαραγγελίας
    supplier: str | None = None         # προμηθευτής
    location: str | None = None         # θέση/ράφι
    batch: str | None = None            # παρτίδα
    expiry: str | None = None           # λήξη (YYYY-MM-DD)
    barcodes: list[str] = Field(default_factory=list)     # εναλλακτικά barcodes (κύριο = barcode)
    variants: list[dict] = Field(default_factory=list)    # εκδοχές: [{color, size, barcode, stock_qty}]
    cat1_id: str | None = None    # κατηγορία e-shop επιπέδου 1 (υποχρεωτική για for_sale)
    cat2_id: str | None = None    # επιπέδου 2 (γονική = cat1)
    cat3_id: str | None = None    # επιπέδου 3 (γονική = cat2)
    # ── καθαρά ΠΩΛΗΣΙΑΚΑ χαρακτηριστικά e-shop (όχι δεδομένα αποθήκης) ──
    sale_starts_at: datetime | None = None   # flash προσφορά: παράθυρο ισχύος της έκπτωσης (προαιρετικό)
    sale_ends_at: datetime | None = None
    highlights: list[str] = Field(default_factory=list)         # σημεία πώλησης (bullets) στη σελίδα προϊόντος
    related_barcodes: list[str] = Field(default_factory=list)   # cross-sell «συχνά μαζί»
    points_multiplier: float = Field(1.0, ge=1, le=10)          # bonus πόντοι πιστότητας για το είδος (×)


@router.post("")
async def upsert_product(body: ProductIn, ctx: TenantContext = Depends(require(_PERM, module=_MODULE))):
    return await _repo(ctx).upsert(body.model_dump())


@router.delete("/{barcode}")
async def delete_product(barcode: str, ctx: TenantContext = Depends(require(_PERM, module=_MODULE))):
    return await _repo(ctx).delete(barcode)


# ── ΑΠΟΘΗΚΗ (master inventory + κινήσεις αποθέματος) ─────────────────────────
def _mrepo(ctx: TenantContext) -> StockMovementRepository:
    return StockMovementRepository(tenant_id=ctx.tenant_id)


@router.get("/warehouse")
async def warehouse(q: str = "", type: str | None = None, low_stock: bool = False,
                    expiring: bool = False, include_inactive: bool = True, page: int = 1,
                    page_size: int = 100,
                    cat1: str | None = None, cat2: str | None = None, cat3: str | None = None,
                    for_sale: bool | None = None, stock: str | None = None,
                    supplier: str | None = None, no_image: bool = False, no_category: bool = False,
                    ctx: TenantContext = Depends(require(_PERM, module=_MODULE))):
    repo = _repo(ctx)
    return {**await repo.warehouse(q=q, ptype=type, low_stock=low_stock, expiring=expiring,
                                   include_inactive=include_inactive, cat1=cat1, cat2=cat2, cat3=cat3,
                                   for_sale=for_sale, stock=stock, supplier=supplier,
                                   no_image=no_image, no_category=no_category,
                                   page=page, page_size=page_size),
            "summary": await repo.warehouse_summary(),
            "suppliers": await repo.warehouse_suppliers()}


class FlagsIn(BaseModel):
    barcode: str
    for_sale: bool | None = None
    active: bool | None = None


@router.post("/warehouse/flags")
async def set_flags(body: FlagsIn, ctx: TenantContext = Depends(require(_PERM, module=_MODULE))):
    return await _repo(ctx).set_flags(body.barcode, for_sale=body.for_sale, active=body.active)


class MoveIn(BaseModel):
    barcode: str
    kind: str = "in"                    # in=παραλαβή · out=πώληση/εξαγωγή · adjust=απογραφή · waste=απόσυρση
    qty: int = Field(1, ge=1)
    reason: str | None = None
    batch: str | None = None
    expiry: str | None = None
    cost_cents: int | None = None       # κόστος παραλαβής (ενημερώνει τη χονδρική)
    set_to: int | None = None           # απογραφή: όρισε ΑΠΟΛΥΤΟ υπόλοιπο (αντί delta)


@router.post("/warehouse/move")
async def stock_move(body: MoveIn, ctx: TenantContext = Depends(require(_PERM, module=_MODULE))):
    """Κίνηση αποθέματος: εφαρμόζει στο είδος + καταγράφει στο ledger (audit trail)."""
    repo = _repo(ctx)
    delta = None
    if body.set_to is None:
        delta = body.qty if body.kind == "in" else -body.qty   # in=+ · out/waste/adjust(delta)=−
    new_stock = await repo.apply_stock(body.barcode, delta=delta, set_to=body.set_to,
                                       expiry=body.expiry, batch=body.batch, cost_cents=body.cost_cents)
    if new_stock is None:
        raise HTTPException(status.HTTP_404_NOT_FOUND, detail={"error": "product_not_found"})
    await _mrepo(ctx).add(barcode=body.barcode, kind=body.kind, qty=(body.qty if body.set_to is None else new_stock),
                          reason=body.reason or "", batch=body.batch or "", expiry=body.expiry or "",
                          cost_cents=body.cost_cents, by=getattr(ctx, "email", None), new_stock=new_stock)
    return {"ok": True, "stock_qty": new_stock}


@router.get("/warehouse/{barcode}/movements")
async def stock_history(barcode: str, ctx: TenantContext = Depends(require(_PERM, module=_MODULE))):
    return {"items": await _mrepo(ctx).history(barcode)}


# ── ΕΥΕΛΙΚΤΗ ΕΙΣΑΓΩΓΗ Excel/CSV (self-service column mapping) ─────────────────
async def _read_capped(file: UploadFile) -> bytes:
    content = await file.read(_MAX_XML + 1)
    if len(content) > _MAX_XML:
        raise HTTPException(status.HTTP_413_REQUEST_ENTITY_TOO_LARGE,
                            detail={"error": "file_too_large", "max_bytes": _MAX_XML})
    return content


@router.post("/import/preview")
async def import_preview(file: UploadFile = File(...),
                         ctx: TenantContext = Depends(require(_PERM, module=_MODULE))):
    """Διαβάζει το αρχείο & επιστρέφει ΔΕΙΓΜΑ (πρώτες γραμμές × στήλες) — ο χρήστης βλέπει τι έχει πού
    και ορίζει mapping. Δεν υποθέτουμε επικεφαλίδες/θέσεις (ο καθένας έχει διαφορετικό Excel)."""
    from app.repositories.pharmacy_catalog import parse_spreadsheet
    rows = parse_spreadsheet(await _read_capped(file), file.filename or "")
    cols = max((len(r) for r in rows[:100]), default=0)
    return {"columns": cols, "total_rows": len(rows),
            "rows": [(r + [""] * (cols - len(r)))[:cols] for r in rows[:20]]}


@router.post("/import/commit")
async def import_commit(file: UploadFile = File(...), mapping: str = Form(...),
                        start_row: int = Form(1), default_type: str = Form("parapharmacy"),
                        for_sale: bool = Form(False),
                        ctx: TenantContext = Depends(require(_PERM, module=_MODULE))):
    """Εισάγει με βάση το mapping ({field: col_index}) & την αρχική γραμμή που όρισε ο χρήστης."""
    from app.repositories.pharmacy_catalog import parse_spreadsheet
    try:
        m = json.loads(mapping) if mapping else {}
    except json.JSONDecodeError:
        raise HTTPException(status.HTTP_400_BAD_REQUEST, detail={"error": "bad_mapping"})
    if not isinstance(m, dict) or m.get("barcode") in (None, "") or m.get("name") in (None, ""):
        raise HTTPException(status.HTTP_400_BAD_REQUEST,
                            detail={"error": "barcode_name_required"})
    rows = parse_spreadsheet(await _read_capped(file), file.filename or "")
    return await _repo(ctx).import_mapped(rows, m, start_row=start_row,
                                          default_type=default_type, for_sale=for_sale)


@router.post("/import-xml")
async def import_xml(file: UploadFile = File(...), row_tag: str = Form(...),
                     mapping: str = Form(...), default_type: str = Form("parapharmacy"),
                     ctx: TenantContext = Depends(require(_PERM, module=_MODULE))):
    """Upload the commercial program's XML + a field mapping (JSON) → upsert products by barcode."""
    try:
        m = json.loads(mapping)
    except json.JSONDecodeError:
        return {"ok": False, "error": "bad_mapping_json"}
    ctype = (file.content_type or "").split(";")[0].strip().lower()
    if ctype not in _XML_CTYPES:
        raise HTTPException(status.HTTP_415_UNSUPPORTED_MEDIA_TYPE,
                            detail={"error": "unsupported_type", "content_type": ctype})
    content = await file.read(_MAX_XML + 1)   # read at most cap+1 → bounded memory
    if len(content) > _MAX_XML:
        raise HTTPException(status.HTTP_413_REQUEST_ENTITY_TOO_LARGE,
                            detail={"error": "file_too_large", "max_bytes": _MAX_XML})
    return await _repo(ctx).import_xml(content, row_tag=row_tag, mapping=m, default_type=default_type)


# ── Εκπτωτικές καμπάνιες (έκπτωση σε ΟΜΑΔΑ ειδών: κατηγορίες ή/και ετικέτες) ──────────────
# ΚΑΝΟΝΑΣ: τα συνταγογραφούμενα ΔΕΝ παίρνουν ΠΟΤΕ έκπτωση καμπάνιας — επιβάλλεται
# server-side στη μηχανή τιμολόγησης (orders_delivery.create_order → campaign_pct_for).
class CampaignIn(BaseModel):
    id: str | None = None
    name: str = Field(..., min_length=1, max_length=120)
    discount_pct: int = Field(..., ge=1, le=90)
    categories: list[str] = []
    tags: list[str] = []
    cat_ids: list[str] = []      # στόχευση σε κόμβους δέντρου κατηγοριών e-shop (π.χ. «Αντιγήρανση»)
    barcodes: list[str] = []     # στόχευση σε συγκεκριμένα είδη
    active: bool = True
    starts_at: datetime | None = None
    ends_at: datetime | None = None


def _crepo(ctx: TenantContext) -> ShopCampaignRepository:
    return ShopCampaignRepository(tenant_id=ctx.tenant_id)


@router.get("/campaigns")
async def list_campaigns(ctx: TenantContext = Depends(require(_PERM, module=_MODULE))):
    return {"items": await _crepo(ctx).list()}


@router.post("/campaigns")
async def save_campaign(body: CampaignIn, ctx: TenantContext = Depends(require(_PERM, module=_MODULE))):
    data = body.model_dump()
    data["_id"] = data.pop("id", None)
    return await _crepo(ctx).upsert(data)


@router.delete("/campaigns/{campaign_id}")
async def delete_campaign(campaign_id: str, ctx: TenantContext = Depends(require(_PERM, module=_MODULE))):
    return await _crepo(ctx).delete(campaign_id)


# ── ΑΥΤΟΜΑΤΕΣ προσφορές καλαθιού (Shopify-style): έκπτωση σε όλη την παραγγελία + δωρεάν μεταφορικά ──
# ΚΑΝΟΝΑΣ: η order-έκπτωση εφαρμόζεται ΜΟΝΟ στα μη-συνταγογραφούμενα (server-side, create_order).
class OrderDiscountIn(BaseModel):
    id: str | None = None
    name: str = Field(..., min_length=1, max_length=120)
    discount_type: str = "order"       # order | free_shipping
    value_type: str = "pct"            # pct | fixed
    value: int = Field(1, ge=1)        # pct(1-90) ή cents (fixed)
    min_cents: int = Field(0, ge=0)
    min_qty: int = Field(0, ge=0)
    usage_limit: int = Field(0, ge=0)
    active: bool = True
    starts_at: datetime | None = None
    ends_at: datetime | None = None


def _odrepo(ctx: TenantContext) -> ShopOrderDiscountRepository:
    return ShopOrderDiscountRepository(tenant_id=ctx.tenant_id)


@router.get("/order-discounts")
async def list_order_discounts(ctx: TenantContext = Depends(require(_PERM, module=_MODULE))):
    return {"items": await _odrepo(ctx).list()}


@router.post("/order-discounts")
async def save_order_discount(body: OrderDiscountIn, ctx: TenantContext = Depends(require(_PERM, module=_MODULE))):
    data = body.model_dump()
    data["_id"] = data.pop("id", None)
    return await _odrepo(ctx).upsert(data)


@router.delete("/order-discounts/{did}")
async def delete_order_discount(did: str, ctx: TenantContext = Depends(require(_PERM, module=_MODULE))):
    return await _odrepo(ctx).delete(did)


# ── Προμηθευτής Profarm: αντιστοίχιση barcode → επίσημη φωτογραφία (back-office, staff-only) ──
class ProfarmCredsIn(BaseModel):
    username: str
    password: str = ""


def _sup(ctx: TenantContext) -> SupplierSettingsRepository:
    return SupplierSettingsRepository(tenant_id=ctx.tenant_id)


@router.get("/supplier/profarm")
async def profarm_status(ctx: TenantContext = Depends(require(_PERM, module=_MODULE))):
    return await _sup(ctx).get_profarm()


@router.post("/supplier/profarm")
async def profarm_save(body: ProfarmCredsIn, ctx: TenantContext = Depends(require(_PERM, module=_MODULE))):
    return await _sup(ctx).save_profarm(body.username, body.password)


@router.delete("/supplier/profarm")
async def profarm_delete(ctx: TenantContext = Depends(require(_PERM, module=_MODULE))):
    return await _sup(ctx).delete_profarm()


@router.post("/supplier/profarm/test")
async def profarm_test(ctx: TenantContext = Depends(require(_PERM, module=_MODULE))):
    from app.services import profarm_service
    creds = await _sup(ctx).profarm_creds()
    if not creds:
        return {"ok": False, "error": "not_configured"}
    return {"ok": await profarm_service.test_login(creds["username"], creds["password"])}


@router.get("/supplier/profarm/sync-status")
async def profarm_sync_status(ctx: TenantContext = Depends(require(_PERM, module=_MODULE))):
    from app.services import profarm_service
    return await profarm_service.sync_status(ctx.tenant_id)


@router.post("/supplier/profarm/sync")
async def profarm_sync(batch: int = 25, only_for_sale: bool = False,
                       ctx: TenantContext = Depends(require(_PERM, module=_MODULE))):
    """Μία παρτίδα harvest (ο frontend καλεί επαναληπτικά μέχρι remaining=0 δείχνοντας πρόοδο)."""
    from app.services import profarm_service
    return await profarm_service.sync_batch(ctx.tenant_id, batch=max(1, min(50, batch)), only_for_sale=only_for_sale)


@router.post("/supplier/profarm/sync-stop")
async def profarm_sync_stop(stopped: bool = True,
                            ctx: TenantContext = Depends(require(_PERM, module=_MODULE))):
    """Οριστική διακοπή (stopped=true) ή επανενεργοποίηση (stopped=false) της φωτο-σάρωσης."""
    from app.services import profarm_service
    return await profarm_service.set_sync_stopped(ctx.tenant_id, stopped)


# ── Εισαγωγή ΟΛΟΚΛΗΡΩΝ προϊόντων OTC/παραφαρμάκων από Profarm (δημιουργία νέων + update υπαρχόντων) ──
@router.get("/supplier/profarm/import-status")
async def profarm_import_status(ctx: TenantContext = Depends(require(_PERM, module=_MODULE))):
    from app.services import profarm_service
    return await profarm_service.import_status(ctx.tenant_id)


@router.post("/supplier/profarm/import")
async def profarm_import_start(ctx: TenantContext = Depends(require(_PERM, module=_MODULE))):
    """Ξεκινά την εισαγωγή (το Celery beat συνεχίζει ήπια στο παρασκήνιο)."""
    from app.services import profarm_service
    return await profarm_service.import_chunk(ctx.tenant_id)


@router.delete("/supplier/profarm/import")
async def profarm_import_reset(ctx: TenantContext = Depends(require(_PERM, module=_MODULE))):
    from app.services import profarm_service
    return await profarm_service.import_reset(ctx.tenant_id)


@router.post("/supplier/profarm/classify")
async def profarm_classify(limit: int = 300, ctx: TenantContext = Depends(require(_PERM, module=_MODULE))):
    """AI-ταξινόμηση (haiku) εισαγμένων Profarm προϊόντων χωρίς κατηγορία, από το όνομα."""
    from app.services import profarm_service
    return await profarm_service.classify_new_products(ctx.tenant_id, limit=max(1, min(500, limit)))


# ── Κουπόνια έκπτωσης ────────────────────────────────────────────────────────────────────
# Ισχύουν ΜΟΝΟ πάνω στην αξία των μη-συνταγογραφούμενων ειδών (server-side).
class CouponIn(BaseModel):
    id: str | None = None
    code: str = Field(..., min_length=3, max_length=32)
    kind: str = Field("pct", pattern="^(pct|amount)$")
    value: int = Field(..., ge=1)
    min_order_cents: int = Field(0, ge=0)
    max_uses: int = Field(0, ge=0)          # 0 = απεριόριστο
    expires_at: datetime | None = None
    active: bool = True


@router.get("/coupons")
async def list_coupons(ctx: TenantContext = Depends(require(_PERM, module=_MODULE))):
    return {"items": await ShopCouponRepository(tenant_id=ctx.tenant_id).list()}


@router.post("/coupons")
async def save_coupon(body: CouponIn, ctx: TenantContext = Depends(require(_PERM, module=_MODULE))):
    data = body.model_dump()
    data["_id"] = data.pop("id", None)
    return await ShopCouponRepository(tenant_id=ctx.tenant_id).upsert(data)


@router.delete("/coupons/{coupon_id}")
async def delete_coupon(coupon_id: str, ctx: TenantContext = Depends(require(_PERM, module=_MODULE))):
    return await ShopCouponRepository(tenant_id=ctx.tenant_id).delete(coupon_id)


# ── Πακέτα (bundles): «2+1» ή combo ──────────────────────────────────────────────────────
class BundleLineIn(BaseModel):
    barcode: str
    qty: int = Field(1, ge=1)


class BundleIn(BaseModel):
    id: str | None = None
    name: str = Field(..., min_length=1, max_length=120)
    kind: str = Field("combo", pattern="^(combo|nplusm)$")
    active: bool = True
    # nplusm
    barcode: str | None = None
    buy_qty: int = Field(2, ge=1)
    free_qty: int = Field(1, ge=1)
    # combo
    lines: list[BundleLineIn] = []
    discount_pct: int = Field(10, ge=1, le=90)


@router.get("/bundles")
async def list_bundles(ctx: TenantContext = Depends(require(_PERM, module=_MODULE))):
    return {"items": await ShopBundleRepository(tenant_id=ctx.tenant_id).list()}


@router.post("/bundles")
async def save_bundle(body: BundleIn, ctx: TenantContext = Depends(require(_PERM, module=_MODULE))):
    data = body.model_dump()
    data["_id"] = data.pop("id", None)
    return await ShopBundleRepository(tenant_id=ctx.tenant_id).upsert(data)


@router.delete("/bundles/{bundle_id}")
async def delete_bundle(bundle_id: str, ctx: TenantContext = Depends(require(_PERM, module=_MODULE))):
    return await ShopBundleRepository(tenant_id=ctx.tenant_id).delete(bundle_id)


# ── Προσφορές ΥΠΗΡΕΣΙΩΝ (π.χ. «−30% σπιρομέτρηση») — φαίνονται στο κύκλωμα «Προσφορές» της πύλης.
# Δεν μπαίνουν στο καλάθι· ο πελάτης τις «κλείνει» ως ραντεβού (reuse appointments).
class ServiceOfferIn(BaseModel):
    id: str | None = None
    title: str = Field(..., min_length=2, max_length=120)
    description: str | None = Field(None, max_length=600)
    photo_url: str | None = None
    image_id: str | None = None
    is_free: bool = False
    price_cents: int = Field(0, ge=0)       # «τώρα» (0 αν δωρεάν)
    compare_cents: int = Field(0, ge=0)     # «πριν» (προαιρετικό)
    cta: str = Field("reserve", pattern="^(reserve|info)$")
    active: bool = True
    starts_at: datetime | None = None
    ends_at: datetime | None = None


def _sorepo(ctx: TenantContext) -> ShopServiceOffersRepository:
    return ShopServiceOffersRepository(tenant_id=ctx.tenant_id)


@router.get("/service-offers")
async def list_service_offers(ctx: TenantContext = Depends(require(_PERM, module=_MODULE))):
    return {"items": await _sorepo(ctx).list()}


@router.post("/service-offers")
async def save_service_offer(body: ServiceOfferIn, ctx: TenantContext = Depends(require(_PERM, module=_MODULE))):
    data = body.model_dump()
    data["_id"] = data.pop("id", None)
    return await _sorepo(ctx).upsert(data)


@router.delete("/service-offers/{offer_id}")
async def delete_service_offer(offer_id: str, ctx: TenantContext = Depends(require(_PERM, module=_MODULE))):
    return await _sorepo(ctx).delete(offer_id)
