"""Excel (.xlsx) import of insured-persons' contact details — parsing + template.

The pharmacist exports their customer list from the commercial program and uploads it here;
rows are matched to existing patients by ΑΜΚΑ and their contact details are enriched
(see PatientContactRepository.import_insured). Header detection is accent/case-insensitive.
"""

from __future__ import annotations

import io
import unicodedata

from openpyxl import Workbook, load_workbook

# normalized header → contact field
_ALIASES: dict[str, str] = {
    "amka": "amka", "αμκα": "amka",
    "ονοματεπωνυμο": "full_name", "ονομα": "full_name", "name": "full_name",
    "fullname": "full_name", "επωνυμοονομα": "full_name", "ονομαεπωνυμο": "full_name",
    "κινητο": "mobile", "mobile": "mobile", "cell": "mobile", "κινητοτηλεφωνο": "mobile",
    "τηλεφωνο": "phone", "phone": "phone", "σταθερο": "phone", "landline": "phone", "τηλ": "phone",
    "email": "email", "emailaddress": "email", "ηλεκτρονικοταχυδρομειο": "email", "mail": "email",
    "διευθυνση": "address", "address": "address", "διευθ": "address",
    "πολη": "city", "city": "city",
    "τκ": "postal_code", "ταχυδρομικοςκωδικας": "postal_code", "ταχυδρομικοςκωδικος": "postal_code",
    "postalcode": "postal_code", "zip": "postal_code", "postal": "postal_code",
    "σημειωσεις": "notes", "notes": "notes", "παρατηρησεις": "notes", "σχολια": "notes",
    "συγκαταθεσημαρκετινγκ": "marketing_consent", "συγκαταθεση": "marketing_consent",
    "marketing": "marketing_consent", "marketingconsent": "marketing_consent", "συναινεση": "marketing_consent",
}

TEMPLATE_HEADERS = ["ΑΜΚΑ", "Ονοματεπώνυμο", "Κινητό", "Τηλέφωνο", "Email",
                    "Διεύθυνση", "Πόλη", "ΤΚ", "Σημειώσεις", "Συγκατάθεση μάρκετινγκ"]

_TRUE = {"1", "true", "ναι", "nai", "yes", "y", "x", "✓", "true.", "on"}


def _norm(s) -> str:
    if s is None:
        return ""
    s = str(s).strip().lower()
    s = "".join(c for c in unicodedata.normalize("NFD", s) if unicodedata.category(c) != "Mn")
    return "".join(ch for ch in s if ch.isalnum())


def parse_contacts_xlsx(data: bytes) -> tuple[list[dict], str | None]:
    """(rows, error). Each row holds only present fields; `amka` is digits-only (kept as string)."""
    try:
        wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    except Exception:
        return [], "Μη έγκυρο αρχείο Excel (.xlsx)."
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    header = next(rows_iter, None)
    if header is None:
        wb.close()
        return [], "Άδειο αρχείο."
    colmap: dict[int, str] = {}
    for i, h in enumerate(header):
        field = _ALIASES.get(_norm(h))
        if field and field not in colmap.values():
            colmap[i] = field
    if "amka" not in colmap.values():
        wb.close()
        return [], "Δεν βρέθηκε στήλη «ΑΜΚΑ». Κατέβασε και χρησιμοποίησε το πρότυπο."
    out: list[dict] = []
    for r in rows_iter:
        if r is None:
            continue
        rec: dict = {}
        for i, field in colmap.items():
            if i >= len(r) or r[i] is None:
                continue
            val = r[i]
            if field == "amka":
                digits = "".join(ch for ch in str(val).split(".")[0] if ch.isdigit())
                if digits:
                    rec["amka"] = digits
            elif field == "marketing_consent":
                rec[field] = _norm(val) in {_norm(t) for t in _TRUE}
            else:
                sval = str(val).strip()
                if sval:
                    rec[field] = sval
        if rec.get("amka"):
            out.append(rec)
    wb.close()
    return out, None


def build_template_xlsx() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Ασφαλισμένοι"
    ws.append(TEMPLATE_HEADERS)
    ws.append(["12345678901", "ΠΑΠΑΔΟΠΟΥΛΟΣ ΓΕΩΡΓΙΟΣ", "6971234567", "2101234567",
               "g.papadopoulos@example.gr", "Ερμού 10", "Αθήνα", "10563", "VIP πελάτης", "Ναι"])
    for col, _ in enumerate(TEMPLATE_HEADERS, start=1):
        ws.column_dimensions[chr(64 + col)].width = 22
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
